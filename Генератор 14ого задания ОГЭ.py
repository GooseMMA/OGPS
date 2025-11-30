import os
import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# === ЗАГОЛОВОК ===
print("GooSenor | Генератор заданий ОГЭ №14 | v1.4")
print("=" * 60)

# === Конфигурация ===
BASE_DIR = "generated_tasks"
os.makedirs(BASE_DIR, exist_ok=True)
N_ROWS = 1000

# === Источники данных ===
DISTRICTS = ["С", "В", "Ю", "З", "СВ", "СЗ", "ЮВ", "ЮЗ", "Ц", "Зел", "Центральный", "Заречный", "Подгорный", "Майский", "Новый", "Светлый", "Яснево", "Кировский"]
SUBJECTS = ["математика", "русский язык", "физика", "информатика", "биология", "химия", "английский язык", "история", "обществознание", "география", "физкультура", "литература", "ИЗО", "французский язык", "немецкий язык"]
COUNTRIES = ["РОССИЯ", "ЕГИПЕТ", "ИТАЛИЯ", "ГЕРМАНИЯ", "ФРАНЦИЯ", "ИСПАНИЯ", "ВЕЛИКОБРИТАНИЯ", "США", "УКРАИНА", "КАНАДА", "ТУРЦИЯ", "АВСТРИЯ", "БЕЛЬГИЯ", "АЛЖИР", "АНГЛИЯ", "УЭЛЬС", "ШВЕЙЦАРИЯ", "АВСТРАЛИЯ", "ОАЭ", "БЕЛОРУССИЯ"]
WIND_DIRS = ["С", "СВ", "В", "ЮВ", "Ю", "ЮЗ", "З", "СЗ"]
PRODUCTS = ["яйцо целое", "молоко", "хлеб", "сыр", "яблоко", "говядина", "рис", "картофель", "курица", "рыба", "макароны", "банан", "апельсин", "шоколад", "кефир"]
STATUSES = ["студент", "пенсионер", "служащий"]
CITIES = ["Москва", "Париж", "Каир", "Берлин", "Лондон", "Нью-Йорк"]

EUROPE = ["ИТАЛИЯ", "ГЕРМАНИЯ", "ФРАНЦИЯ", "ИСПАНИЯ", "БЕЛЬГИЯ", "АВСТРИЯ", "ВЕЛИКОБРИТАНИЯ", "АНГЛИЯ", "ПОРТУГАЛИЯ", "ШВЕЙЦАРИЯ"]

# === Вспомогательные функции ===
def save_files(data, questions, formulas, answers, diag_text, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    for col in range(1, len(data[0]) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
    wb.save(os.path.join(output_dir, "таблица_ученика.xlsx"))

    task_text = f"""ЗАДАНИЕ №14

На основании данных в файле «таблица_ученика.xlsx» ответьте на вопросы:

1. {questions[0]}
2. {questions[1]}

{diag_text}
"""
    with open(os.path.join(output_dir, "задание_ученика.txt"), "w", encoding="utf-8") as f:
        f.write(task_text)

    formulas_text = f"""ФОРМУЛЫ И ОТВЕТЫ ДЛЯ ПРОВЕРКИ

Вопрос 1: {questions[0]}
Формула: {formulas[0]}
Ответ: {answers[0]}

Вопрос 2: {questions[1]}
Формула: {formulas[1]}
Ответ: {answers[1]}
"""
    with open(os.path.join(output_dir, "формулы_и_ответы.txt"), "w", encoding="utf-8") as f:
        f.write(formulas_text)

def get_count():
    while True:
        n = input("➤ Сколько заданий сгенерировать? (0 — в меню): ").strip()
        if n in ("", "0"):
            return 0
        if n.isdigit() and int(n) > 0:
            return int(n)
        print("❌ Введите положительное число или 0.")

def choose_type():
    print("\n📊 Выберите тип задания:")
    types = {
        1: "Ученики: округ + предмет + балл",
        2: "Оценки по предметам",
        3: "Погода",
        4: "Города и население",
        5: "Пищевая ценность продуктов",
        6: "Тестирование: пол + возраст + статус",
        7: "Перелёты: город + страна + время"
    }
    for k, v in types.items():
        print(f"  {k}. {v}")
    print("  0. Назад в меню")
    while True:
        choice = input("➤ Ваш выбор: ").strip()
        if choice == "0":
            return None
        if choice.isdigit() and 1 <= int(choice) <= 7:
            return int(choice)
        print("❌ Введите число от 1 до 7 или 0.")

# === Генераторы с вариативностью ===

def generate_type1():
    data = [["округ", "фамилия", "предмет", "балл"]]
    values = []
    for i in range(1, N_ROWS + 1):
        d = random.choice(DISTRICTS)
        s = random.choice(SUBJECTS)
        b = random.randint(200, 1000)
        data.append([d, f"Ученик {i}", s, b])
        values.append((d, s, b))

    # Вариативный порог
    threshold = random.choice([300, 400, 500, 600, 700])
    condition = random.choice([">", ">=", "<"])
    d1 = random.choice(DISTRICTS)
    s1 = random.choice(SUBJECTS)

    if condition == ">":
        ans1 = sum(1 for d, s, b in values if d == d1 and b > threshold)
        q1 = f"Сколько учеников из округа «{d1}» набрали более {threshold} баллов по любому предмету?"
        f1 = f'=СЧЁТЕСЛИМН(A2:A{N_ROWS+1};"{d1}";D2:D{N_ROWS+1};">{threshold}")'
    elif condition == ">=":
        ans1 = sum(1 for d, s, b in values if s == s1 and b >= threshold)
        q1 = f"Сколько учеников выбрали «{s1}» и получили не менее {threshold} баллов?"
        f1 = f'=СЧЁТЕСЛИМН(C2:C{N_ROWS+1};"{s1}";D2:D{N_ROWS+1};">={threshold}")'
    else:  # "<"
        ans1 = sum(1 for d, s, b in values if d == d1 and s == s1 and b < threshold)
        q1 = f"Сколько учеников из округа «{d1}» выбрали «{s1}» и получили менее {threshold} баллов?"
        f1 = f'=СЧЁТЕСЛИМН(A2:A{N_ROWS+1};"{d1}";C2:C{N_ROWS+1};"{s1}";D2:D{N_ROWS+1};"<{threshold}")'

    d2 = random.choice(DISTRICTS)
    scores = [b for d, s, b in values if d == d2]
    ans2 = round(sum(scores) / len(scores), 2) if scores else 0
    q2 = f"Каков средний балл у учеников из округа «{d2}»?"
    f2 = f'=ОКРУГЛ(СРЗНАЧЕСЛИ(A2:A{N_ROWS+1};"{d2}";D2:D{N_ROWS+1});2)'

    diag = f"Постройте круговую диаграмму по округам: «{'», «'.join(random.sample(DISTRICTS, 3))}»."
    return data, [q1, q2], [f1, f2], [ans1, ans2], diag

def generate_type2():
    data = [["Фамилия", "Имя", "Алгебра", "Русский", "Физика", "Информатика"]]
    values = []
    for i in range(1, N_ROWS + 1):
        row = [f"Ученик{i}", f"Ученик{i}"] + [random.randint(2, 5) for _ in range(4)]
        data.append(row)
        values.append(row[2:])

    threshold = random.choice([3, 4, 5])
    condition = random.choice(["не ниже", "выше", "ровно"])
    op = ">=" if condition == "не ниже" else ">" if condition == "выше" else "="
    text_cond = f"{condition} {threshold}" if condition != "ровно" else f"ровно {threshold}"

    subj = random.choice(["Алгебра", "Русский", "Физика", "Информатика"])
    idx = ["Алгебра", "Русский", "Физика", "Информатика"].index(subj)
    col = ["C", "D", "E", "F"][idx]

    # Вариант: по одному предмету
    ans1 = sum(1 for r in values if (r[idx] >= threshold if op == ">=" else r[idx] > threshold if op == ">" else r[idx] == threshold))
    q1 = f"Сколько учеников получили оценку {text_cond} по {subj.lower()}?"
    f1 = f'=СЧЁТЕСЛИ({col}2:{col}{N_ROWS+1};"{op}{threshold}")'

    # Средний по тем, у кого все >= 3 или >=4
    min_grade = random.choice([3, 4])
    qualified = [r for r in values if all(x >= min_grade for x in r)]
    scores2 = [r[idx] for r in qualified]
    ans2 = round(sum(scores2) / len(scores2), 2) if scores2 else 0
    q2 = f"Какой средний балл по {subj.lower()} у учеников, у которых по всем предметам стоят оценки не ниже {min_grade}?"
    f2 = f'=ОКРУГЛ(СРЗНАЧЕСЛИМН({col}2:{col}{N_ROWS+1};C2:C{N_ROWS+1};">={min_grade}";D2:D{N_ROWS+1};">={min_grade}";E2:E{N_ROWS+1};">={min_grade}";F2:F{N_ROWS+1};">={min_grade}");2)'

    diag = f"Постройте круговую диаграмму по оценкам по {subj.lower()} («3», «4», «5»)."
    return data, [q1, q2], [f1, f2], [ans1, ans2], diag

def generate_type3():
    data = [["Дата", "Температура", "Осадки", "Давление", "Ветер", "Скорость ветра"]]
    values = []
    months = ["января", "февраля", "марта", "апреля", "мая", "июня", "июля", "августа", "сентября", "октября", "ноября", "декабря"]
    for i in range(1, N_ROWS + 1):
        day = f"{i % 28 + 1} {random.choice(months)}"
        temp = round(random.uniform(-20, 35), 1)
        precip = round(random.uniform(0, 20), 1)
        pressure = random.randint(730, 770)
        wind = random.choice(WIND_DIRS)
        speed = round(random.uniform(0, 15), 1)
        data.append([day, temp, precip, pressure, wind, speed])
        values.append((temp, precip, wind, speed, day))

    # Вариант 1: по ветру
    if random.choice([True, False]):
        wind1 = random.choice(WIND_DIRS)
        ans1 = sum(1 for temp, precip, wind, speed, day in values if wind == wind1)
        q1 = f"Сколько дней дул ветер с направления «{wind1}»?"
        f1 = f'=СЧЁТЕСЛИ(E2:E{N_ROWS+1};"{wind1}")'
    else:
        temp1 = round(random.uniform(10, 25), 1)
        ans1 = sum(1 for temp, precip, wind, speed, day in values if temp > temp1 and precip == 0)
        q1 = f"Сколько дней температура была выше {temp1}°C и осадков не было?"
        f1 = f'=СЧЁТЕСЛИМН(B2:B{N_ROWS+1};">{temp1}";C2:C{N_ROWS+1};"=0")'

    # Средняя скорость ветра в августе
    august_speeds = [speed for temp, precip, wind, speed, day in values if "августа" in day]
    ans2 = round(sum(august_speeds) / len(august_speeds), 2) if august_speeds else 0
    q2 = "Какова средняя скорость ветра в августе?"
    f2 = f'=ОКРУГЛ(СРЗНАЧЕСЛИ(A2:A{N_ROWS+1};"*августа*";F2:F{N_ROWS+1});2)'

    diag = f"Постройте круговую диаграмму по направлениям ветра: «{'», «'.join(random.sample(WIND_DIRS, 3))}»."
    return data, [q1, q2], [f1, f2], [ans1, ans2], diag

def generate_type4():
    data = [["Город", "Население (тыс.)", "Страна"]]
    values = []
    for i in range(1, N_ROWS + 1):
        city = f"Город {i}"
        pop = round(random.uniform(10, 5000), 2)
        country = random.choice(COUNTRIES)
        data.append([city, pop, country])
        values.append((country, pop))

    # Вариант: по стране или Европе
    if random.choice([True, False]):
        c1 = random.choice(COUNTRIES)
        pop1 = random.choice([100, 200, 500, 1000])
        cond = random.choice([">", "<"])
        if cond == ">":
            ans1 = sum(1 for country, pop in values if country == c1 and pop > pop1)
            q1 = f"Сколько городов в стране «{c1}» с населением более {pop1} тыс.?"
            f1 = f'=СЧЁТЕСЛИМН(C2:C{N_ROWS+1};"{c1}";B2:B{N_ROWS+1};">{pop1}")'
        else:
            ans1 = sum(1 for country, pop in values if country == c1 and pop < pop1)
            q1 = f"Сколько городов в стране «{c1}» с населением менее {pop1} тыс.?"
            f1 = f'=СЧЁТЕСЛИМН(C2:C{N_ROWS+1};"{c1}";B2:B{N_ROWS+1};"<{pop1}")'
    else:
        europe_list = random.sample(EUROPE, 3)
        pop1 = random.choice([100, 200, 500])
        ans1 = sum(1 for country, pop in values if country in europe_list and pop > pop1)
        europe_str = ", ".join(europe_list)
        q1 = f"Сколько городов в странах Европы ({europe_str}) с населением более {pop1} тыс.?"
        parts = [f'СЧЁТЕСЛИМН(C2:C{N_ROWS+1};"{c}";B2:B{N_ROWS+1};">{pop1}")' for c in europe_list]
        f1 = "=" + " + ".join(parts)

    c2 = random.choice(COUNTRIES)
    pops2 = [pop for country, pop in values if country == c2]
    ans2 = round(sum(pops2) / len(pops2), 2) if pops2 else 0
    q2 = f"Каково среднее население городов в стране «{c2}»?"
    f2 = f'=ОКРУГЛ(СРЗНАЧЕСЛИ(C2:C{N_ROWS+1};"{c2}";B2:B{N_ROWS+1});2)'

    diag = f"Постройте круговую диаграмму по странам: «{'», «'.join(random.sample(COUNTRIES, 3))}»."
    return data, [q1, q2], [f1, f2], [ans1, ans2], diag

def generate_type5():
    data = [["Продукт", "Жиры", "Белки", "Углеводы", "Калорийность"]]
    values = []
    for _ in range(N_ROWS):
        prod = random.choice(PRODUCTS)
        fats = round(random.uniform(0, 40), 1)
        proteins = round(random.uniform(0, 30), 1)
        carbs = round(random.uniform(0, 80), 1)
        kcal = round(9*fats + 4*proteins + 4*carbs, 1)
        data.append([prod, fats, proteins, carbs, kcal])
        values.append((fats, proteins, carbs, kcal))

    # Вариативные условия
    choice = random.choice([
        ("углеводов", "белков", "D", "C", "<", ">"),
        ("жиров", "калорийности", "B", "E", "<", ">"),
        ("белков", "углеводов", "C", "D", ">", "<")
    ])
    carb_col, prot_col, carb_let, prot_let, op1, op2 = choice
    c1 = random.choice([10, 15, 20, 25])
    p1 = random.choice([20, 25, 30])

    if op1 == "<" and op2 == ">":
        ans1 = sum(1 for fats, proteins, carbs, kcal in values if carbs < c1 and proteins > p1)
        q1 = f"Сколько продуктов содержат менее {c1} г {carb_col} и более {p1} г {prot_col}?"
        f1 = f'=СЧЁТЕСЛИМН({carb_let}2:{carb_let}{N_ROWS+1};"<{c1}";{prot_let}2:{prot_let}{N_ROWS+1};">{p1}")'
    elif op1 == "<" and op2 == ">":
        f1_val = random.choice([5, 10, 15])
        kcal1 = random.choice([200, 250, 300])
        ans1 = sum(1 for fats, proteins, carbs, kcal in values if fats < f1_val and kcal > kcal1)
        q1 = f"Сколько продуктов содержат менее {f1_val} г жиров и более {kcal1} Ккал?"
        f1 = f'=СЧЁТЕСЛИМН(B2:B{N_ROWS+1};"<{f1_val}";E2:E{N_ROWS+1};">{kcal1}")'
    else:
        ans1 = sum(1 for fats, proteins, carbs, kcal in values if proteins > p1 and carbs < c1)
        q1 = f"Сколько продуктов содержат более {p1} г белков и менее {c1} г углеводов?"
        f1 = f'=СЧЁТЕСЛИМН(C2:C{N_ROWS+1};">{p1}";D2:D{N_ROWS+1};"<{c1}")'

    low_fat = [kcal for fats, proteins, carbs, kcal in values if fats < 10]
    ans2 = round(sum(low_fat) / len(low_fat), 2) if low_fat else 0
    q2 = "Какова средняя калорийность продуктов с содержанием жиров менее 10 г?"
    f2 = f'=ОКРУГЛ(СРЗНАЧЕСЛИ(B2:B{N_ROWS+1};"<10";E2:E{N_ROWS+1});2)'

    diag = "Постройте круговую диаграмму по среднему содержанию жиров, белков и углеводов в первых 100 продуктах."
    return data, [q1, q2], [f1, f2], [ans1, ans2], diag

def generate_type6():
    data = [["номер участника", "пол", "возраст", "тест 1", "тест 2"]]
    values = []
    for i in range(1, N_ROWS + 1):
        gender = random.choice(["муж", "жен"])
        age = random.randint(15, 80)
        test1 = random.randint(0, 100)
        test2 = random.randint(0, 100)
        data.append([f"участник {i}", gender, age, test1, test2])
        values.append((gender, age, test1))

    # Вариативные условия
    if random.choice([True, False]):
        age1 = random.choice([40, 50, 60])
        score1 = random.choice([40, 50, 60])
        ans1 = sum(1 for gender, age, test1 in values if gender == "жен" and age > age1 and test1 > score1)
        q1 = f"Сколько женщин старше {age1} лет набрали более {score1} баллов на первом тесте?"
        f1 = f'=СЧЁТЕСЛИМН(B2:B{N_ROWS+1};"жен";C2:C{N_ROWS+1};">{age1}";D2:D{N_ROWS+1};">{score1}")'
    else:
        age1 = random.choice([25, 30, 35])
        score1 = random.choice([30, 40, 50])
        ans1 = sum(1 for gender, age, test1 in values if gender == "муж" and age < age1 and test1 < score1)
        q1 = f"Сколько мужчин младше {age1} лет набрали менее {score1} баллов на первом тесте?"
        f1 = f'=СЧЁТЕСЛИМН(B2:B{N_ROWS+1};"муж";C2:C{N_ROWS+1};"<{age1}";D2:D{N_ROWS+1};"<{score1}")'

    young_men = [test1 for gender, age, test1 in values if gender == "муж" and age < 30]
    ans2 = round(sum(young_men) / len(young_men), 2) if young_men else 0
    q2 = "Какой средний балл на первом тесте у мужчин младше 30 лет?"
    f2 = f'=ОКРУГЛ(СРЗНАЧЕСЛИМН(D2:D{N_ROWS+1};B2:B{N_ROWS+1};"муж";C2:C{N_ROWS+1};"<30");2)'

    diag = "Постройте круговую диаграмму по количеству женщин-пенсионеров, женщин-студентов и женщин-служащих."
    return data, [q1, q2], [f1, f2], [ans1, ans2], diag

def generate_type7():
    data = [["Город", "Страна", "Время в пути (ч)"]]
    values = []
    for _ in range(N_ROWS):
        city = random.choice(CITIES)
        country = random.choice(COUNTRIES)
        time = random.randint(2, 12)
        data.append([city, country, time])
        values.append((country, time))

    # Вариант: по Европе или одной стране
    if random.choice([True, False]):
        hours = random.randint(4, 8)
        europe_str = ", ".join(EUROPE)
        ans1 = sum(1 for country, time in values if country in EUROPE and time < hours)
        q1 = f"Сколько перелётов в страны Европы ({europe_str}) длятся менее {hours} часов?"
        parts = [f'СЧЁТЕСЛИМН(B2:B{N_ROWS+1};"{c}";C2:C{N_ROWS+1};"<{hours}")' for c in EUROPE]
        f1 = "=" + " + ".join(parts)
    else:
        c1 = random.choice(COUNTRIES)
        h1 = random.randint(4, 8)
        cond = random.choice([">", "<"])
        if cond == ">":
            ans1 = sum(1 for country, time in values if country == c1 and time > h1)
            q1 = f"Сколько перелётов в страну «{c1}» длятся более {h1} часов?"
            f1 = f'=СЧЁТЕСЛИМН(B2:B{N_ROWS+1};"{c1}";C2:C{N_ROWS+1};">{h1}")'
        else:
            ans1 = sum(1 for country, time in values if country == c1 and time < h1)
            q1 = f"Сколько перелётов в страну «{c1}» длятся менее {h1} часов?"
            f1 = f'=СЧЁТЕСЛИМН(B2:B{N_ROWS+1};"{c1}";C2:C{N_ROWS+1};"<{h1}")'

    c2 = random.choice(COUNTRIES)
    times2 = [time for country, time in values if country == c2]
    ans2 = round(sum(times2) / len(times2), 2) if times2 else 0
    q2 = f"Каково среднее время в пути для перелётов в страну «{c2}»?"
    f2 = f'=ОКРУГЛ(СРЗНАЧЕСЛИ(B2:B{N_ROWS+1};"{c2}";C2:C{N_ROWS+1});2)'

    diag = f"Постройте круговую диаграмму по странам: «{'», «'.join(random.sample(COUNTRIES, 3))}»."
    return data, [q1, q2], [f1, f2], [ans1, ans2], diag

# === Карта типов ===
TYPE_GENERATORS = {
    1: generate_type1,
    2: generate_type2,
    3: generate_type3,
    4: generate_type4,
    5: generate_type5,
    6: generate_type6,
    7: generate_type7
}

# === Основной цикл ===
def main():
    while True:
        print("\n📋 МЕНЮ")
        print("1. Сгенерировать случайное задание")
        print("2. Выбрать тип задания")
        print("3. Выход")
        choice = input("\n➤ Ваш выбор: ").strip()

        if choice == "1":
            count = get_count()
            if count == 0:
                continue
            for i in range(1, count + 1):
                tid = random.randint(1, 7)
                gen_func = TYPE_GENERATORS[tid]
                data, qs, fs, ans, diag = gen_func()
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                folder = f"генерация_{timestamp}_тип{tid}_задание{i}"
                save_files(data, qs, fs, ans, diag, os.path.join(BASE_DIR, folder))
                print(f"✅ {folder}")

        elif choice == "2":
            tid = choose_type()
            if tid is None:
                continue
            count = get_count()
            if count == 0:
                continue
            gen_func = TYPE_GENERATORS[tid]
            for i in range(1, count + 1):
                data, qs, fs, ans, diag = gen_func()
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                folder = f"генерация_{timestamp}_тип{tid}_задание{i}"
                save_files(data, qs, fs, ans, diag, os.path.join(BASE_DIR, folder))
                print(f"✅ {folder}")

        elif choice == "3":
            print("\n👋 До свидания!")
            break

        else:
            print("❌ Неверный выбор.")

if __name__ == "__main__":
    main()